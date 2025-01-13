from copy import deepcopy
import os
import numpy as np
import pandas as pd

# For Lasso regression
from numpy import iterable
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from sklearn.model_selection import train_test_split, GridSearchCV
from sklearn.linear_model import Lasso
from sklearn.multiclass import OneVsRestClassifier

# For OneHotEncoder
from sklearn.preprocessing import OneHotEncoder
from sklearn.compose import ColumnTransformer
from sklearn.model_selection import GridSearchCV, KFold, cross_val_score

# For RandomForestClassifier
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from sklearn.metrics import confusion_matrix, balanced_accuracy_score
from sklearn.model_selection import StratifiedKFold

# For GroupKFold
from sklearn.preprocessing import LabelEncoder
from sklearn.model_selection import GroupKFold

# For GradientBoost
from sklearn.datasets import make_hastie_10_2
from sklearn.ensemble import GradientBoostingClassifier

# For Report (Confusion Matrix/Classification Report/AUROC)
from sklearn import metrics
from sklearn.metrics import ConfusionMatrixDisplay
from sklearn.svm import SVC
from sklearn.metrics import classification_report, f1_score, accuracy_score, average_precision_score
from sklearn.metrics import roc_auc_score, RocCurveDisplay
from sklearn.model_selection import StratifiedKFold
from sklearn.metrics import roc_curve, auc

# For Decision Tree
from sklearn import tree

from mpl_toolkits.axes_grid1.inset_locator import inset_axes
from joblib import dump, Parallel, delayed
from datetime import datetime, timedelta

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


class trained_model:
    def __init__(self) -> None:
        pass

    def five_fold_cross_train(self, self_pip, estimator, type_cv: str = "grouped", grouped_split_on: str = "split_int"):
        X_train, y_train = self_pip.data.X, self_pip.data.y
        models = []
        model_with_info = {}
        iterator = -1

        # construct cross-validation method for inner and outer loop for hyperparameter tuning and 5fold training
        if self_pip.user_input.training["cross_validation_method"] == "grouped":
            gcv = GroupKFold(n_splits=self_pip.user_input.training["n_splits"])
            scv = StratifiedKFold(
                n_splits=self_pip.user_input.training["n_splits"],
                shuffle=True,
                random_state=self_pip.user_input.training["random_state"],
            )
            cv_outer, cv_inner = gcv, scv
            le = LabelEncoder()
            cv_out = cv_outer.split(X_train, y_train, groups=le.fit_transform(X_train[grouped_split_on]))
        elif self_pip.user_input.training["cross_validation_method"] == "stratified":
            scv = StratifiedKFold(
                n_splits=self_pip.user_input.training["n_splits"],
                shuffle=True,
                random_state=self_pip.user_input.training["random_state"],
            )
            cv_outer, cv_inner = scv, scv
            cv_out = cv_outer.split(X_train, y_train)
            le = LabelEncoder()
        # start outer loop:
        ## define cv object
        def run_inner_loop(iterator,train_idx, test_idx,estimator=estimator,models=models,model_with_info=model_with_info,self_pip=self_pip):
            X_train_inner, X_test_inner = X_train.iloc[train_idx], X_train.iloc[test_idx]
            y_train_inner, y_test_inner = y_train.iloc[train_idx], y_train.iloc[test_idx]

            # construct the estimator:
            estimator = estimator # see training.models.py for the estimators

            # tune the hyperparameters in the inner loop:
            model = GridSearchCV(
                estimator=estimator,
                param_grid=self_pip.user_input.training["hyper_para_options"],
                verbose=self_pip.user_input.training["verbose"],
                cv=cv_inner,
                scoring=self_pip.user_input.training["scoring_grid_search"],
                n_jobs=-1 # all processors will be used
            )  # 'roc_auc_ovr_weighted',‘balanced_accuracy’ also an option look at the overfitting
            model.fit(
                self_pip.ohe.transform(X_train_inner), y_train_inner.status
            )  # ,groups=le.fit_transform(X_train_inner['split_int']))
            models.append(model)
            fpr_test, tpr_test, _ = roc_curve(
                y_test_inner, model.predict_proba(self_pip.ohe.transform(X_test_inner))[:, 1]
            )
            fpr_train, tpr_train, _ = roc_curve(
                y_train_inner, model.predict_proba(self_pip.ohe.transform(X_train_inner))[:, 1]
            )
            model_with_info.update(
                {
                    f"model_{iterator}": {
                        "model": model,
                        "X_train_inner": X_train_inner,
                        "X_test_inner": X_test_inner,
                        "y_train_inner": y_train_inner,
                        "y_test_inner": y_test_inner,
                        "y_pred_test": pd.Series(model.predict_proba(self_pip.ohe.transform(X_test_inner))[:, 1],index=X_test_inner.index,name='status'),
                        "y_pred_train": pd.Series(model.predict_proba(self_pip.ohe.transform(X_train_inner))[:, 1],index=X_train_inner.index,name='status'),
                        "perf_on_test": {"fpr": fpr_test, "tpr": tpr_test, "roc_auc_raw": auc(fpr_test, tpr_test)},
                        "perf_on_train": {"fpr": fpr_train, "tpr": tpr_train, "roc_auc_raw": auc(fpr_train, tpr_train)},
                        "user_input": self_pip.user_input,
                    }
                }
            )
            return model_with_info, models
        results=Parallel(n_jobs=-1)(delayed(run_inner_loop)(iterator,train_idx,test_idx) for iterator,(train_idx, test_idx) in enumerate(cv_out))
        for (model_with_i, model) in  results:
            model_with_info.update(model_with_i)
            models.append(model)
        self.model_with_info = model_with_info
        self.models = models

    def roc_test_train(self, key_df, ax=None, interpolate=False):

        tprs = pd.DataFrame()
        aucs = []
        for key, val in self.model_with_info.items():
            if key.startswith("model"):
                fpr = val.get(key_df).get("fpr")
                tpr = val.get(key_df).get("tpr")
                fpr_base = np.linspace(0, 1, 100)
                tpr_interpol = np.interp(fpr_base, fpr, tpr)
                auc_interpol = auc(fpr_base, tpr_interpol)
                if interpolate:
                    ax.plot(fpr_base, tpr_interpol, color="lightgray", lw=2)
                else:
                    ax.plot(fpr, tpr, color="lightgray", lw=2)
                tprs[key_df + "_interpol_tpr_" + key] = tpr_interpol
                aucs.append(auc_interpol)
        if interpolate:
            tprs["mean_tpr"] = tprs.mean(axis=1)
            ax.plot(fpr_base, tprs.mean_tpr, lw=2, label=f"Mean ROCcurve(AUC = {round(np.mean(aucs),2)})")
        ax.legend()
        ax.set_title(key_df)
        self.model_with_info.update({key_df: {"tprs_interpol": tprs, "aucs_interpol": aucs}})


class eval:
    def __init__(self, pip_self, only_val=False):
        self.only_val = only_val
        """Evaluate the performance of the models on the training and testing sets, as well as the performance of the mastermodel on the validation model.


        Args:
            pip_self (pipeline_object): pipeline object of the model
            only_val (bool, optional): If you are only interested on the performance on the validation set you can set is to True (not recommened). Defaults to False.
        """

        if not only_val:
            # generate some print outputs:
            test_, train_ = [], []
            print("\n")
            print("Evaluation of the Model".center(80, "-"))
            for key, val in pip_self.trained_model.model_with_info.items():
                print(key.center(80, "-"))
                test = val["perf_on_test"]["roc_auc_raw"]
                print("auc_test:", test)
                train = val["perf_on_train"]["roc_auc_raw"]
                print("auc_train:", train)
                print("best_params:", val["model"].best_params_)
                print("".center(80, "-"))
                test_.append(test)
                train_.append(train)
            print("ROCcurve AUC".center(80, "-"))
            print("test_mean:", np.mean(test_))
            print("train_mean:", np.mean(train_))

        ### now get the TPRS and AUCS
        tprs_test = pd.DataFrame()
        tprs_train = pd.DataFrame()
        tprs_val = pd.DataFrame()
        aucs_test = []
        aucs_train = []
        aucs_val = []
        auprcs_test = []
        auprcs_train = []
        auprcs_val = []
        ohe = pip_self.ohe

        has_validation_data = (
            hasattr(pip_self.data, "y_val") and pip_self.data.y_val is not None
        )  # Check whether validation data exists
        print(f'The training and testing is evaluated on the {pip_self.user_input.target} labels.')
        print(f'The y_val is now set to the target to validate on: {pip_self.user_input.target_to_validate_on}\nFrom now on this will be used as y_val! change if needed.')
        print(f'Using {pip_self.user_input.target_to_validate_on} for the evaluation on the validation set.')

        for key, val in pip_self.trained_model.model_with_info.items():
            model = val["model"]
            if not only_val:
                # for the testing:
                fpr, tpr, thres = roc_curve(val["y_test_inner"], val["y_pred_test"].tolist())
                fpr_test, tpr_test, thres_test = fpr.copy(), tpr.copy(), thres.copy()
                fpr_base = np.linspace(0, 1, 100)
                tpr = np.interp(fpr_base, fpr, tpr)
                aucs_test.append(auc(fpr_base, tpr))
                tprs_test[key] = tpr
                auprcs_test.append(average_precision_score(val["y_test_inner"], val["y_pred_test"].tolist()))
                # for the training
                fpr, tpr, thres = roc_curve(val["y_train_inner"], val["y_pred_train"].tolist())
                fpr_train, tpr_train, thres_train = fpr.copy(), tpr.copy(), thres.copy()
                fpr_base = np.linspace(0, 1, 100)
                tpr = np.interp(fpr_base, fpr, tpr)
                aucs_train.append(auc(fpr_base, tpr))
                tprs_train[key] = tpr
                auprcs_train.append(average_precision_score(val["y_train_inner"], val["y_pred_train"].tolist()))

            # for the validation
            if has_validation_data:
                # set the y_val to the desired target
                pip_self.data.y_val = pip_self.data.y_val_orig[pip_self.user_input.target_to_validate_on].copy()

                val.update(
                    {
                        "y_val": pip_self.data.y_val_orig[pip_self.user_input.target_to_validate_on],
                        "y_pred_val": pd.Series(model.predict_proba(ohe.transform(pip_self.data.X_val))[:, 1],index=pip_self.data.X_val.index,name='status'),
                    }
                )
                fpr, tpr, thres = roc_curve(val.get("y_val"), val.get("y_pred_val").to_list())
                fpr_val, tpr_val, thres_val = fpr.copy(), tpr.copy(), thres.copy()
                fpr_base = np.linspace(0, 1, 100)
                tpr = np.interp(fpr_base, fpr, tpr)
                aucs_val.append(auc(fpr_base, tpr))
                tprs_val[key] = tpr
                auprcs_val.append(average_precision_score(val.get("y_val"), val.get("y_pred_val").to_list()))
            else:
                raise ValueError("Validation data not available")
        if not only_val:
            self.train = {
                "tprs": tprs_train,
                "aucs": aucs_train,
                "auprcs": auprcs_train,
                "raw": [fpr_train, tpr_train, thres_train],
            }
            self.test = {
                "tprs": tprs_test,
                "aucs": aucs_test,
                "auprcs": auprcs_test,
                "raw": [fpr_test, tpr_test, thres_test],
            }
        if has_validation_data:
            self.val = {"tprs": tprs_val, "aucs": aucs_val, "auprcs": auprcs_val, "raw": [fpr_val, tpr_val, thres_val]}
        else:
            self.val = None
        self.feature_imp = {"PlotX": None}
        self.test_train_pred = {}

    def save_performance_combination(self, pip_self, tprs, pred_values, y_true, true_cancerreg=None, cohort=None):
        """ TODO: rework with inheritance of the function for test, train val!!!!!

        """
        print("Starting save_performance_combination...")
        row_subset = pip_self.user_input.row_subset
        col_subset = pip_self.user_input.col_subset
        subset = pip_self.name
        #####################################- val -
        filename = f"TPRS_combined.xlsx"
        if cohort is None:
            # save the TPRS for plotting together
            combined_output_path = os.path.join(pip_self.pipeline_output_path, "combined_output/")
            os.makedirs(combined_output_path, exist_ok=True)
        else:
            combined_output_path = os.path.join(pip_self.pipeline_output_path, "combined_output/" + cohort + "/")
            if not os.path.exists(combined_output_path):
                os.makedirs(combined_output_path, exist_ok=True)

        tprs_combined_path = os.path.join(combined_output_path, filename)
        if os.path.exists(tprs_combined_path):
            # Load existing TPRS data
            tprs_combined = pd.read_excel(tprs_combined_path)
        else:
            # Initialize an empty DataFrame if no file exists
            tprs_combined = pd.DataFrame()

        # Data from the current model run
        tprs_export = tprs.rename(
            axis=1, mapper=lambda x: f"{pip_self.user_input.row_subset}_{pip_self.user_input.col_subset}_{str(x)}"
        )

        # Identify columns related to the current row and column subset
        current_model_columns = [col for col in tprs_combined.columns if col.startswith(f"{row_subset}_{col_subset}_")]

        # If columns exist for this model configuration, replace them
        if current_model_columns:
            # Drop the old columns from the same model configuration
            tprs_combined = tprs_combined.drop(columns=current_model_columns)

        # Concatenate the new data
        tprs_combined_export = pd.concat([tprs_combined, tprs_export], axis=1)

        # Save the updated DataFrame back to Excel
        tprs_combined_export.to_excel(tprs_combined_path, index=False)
        print(f"TPRS Data exported to {tprs_combined_path}")

        ################# Saving the prediction values##########
        if cohort == "val" and hasattr(pip_self.data, "y_val") and pip_self.data.y_val is not None:
            filename = f"Prediction_values_combined.xlsx"
            full_path = os.path.join(combined_output_path, filename)
            if not os.path.exists(full_path):
                wb = Workbook()
                wb.save(full_path)
                wb.close()


            wb = load_workbook(full_path)
            sheet_name = f"{row_subset}_{col_subset}"
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]  # Delete the sheet if it exists
            wb.save(full_path)
            wb.close()


            # Now append using ExcelWriter without worrying about sheet existing
            with pd.ExcelWriter(full_path, mode="a", engine="openpyxl") as writer:
                pip_self.eval.test_train_pred.get('val').reset_index().to_excel(writer, sheet_name=sheet_name, index=False)

            print(f"Prediction Values Data exported to {full_path}")
        elif cohort == "train" and "train" in pip_self.eval.test_train_pred and pip_self.eval.train is not None:
            filename = f"Prediction_values_combined.xlsx"
            full_path = os.path.join(combined_output_path, filename)
            if not os.path.exists(full_path):
                wb = Workbook()
                wb.save(full_path)
                wb.close()

            wb = load_workbook(full_path)
            sheet_name = f"{row_subset}_{col_subset}"
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]  # Delete the sheet if it exists
            wb.save(full_path)
            wb.close()

            # Now append using ExcelWriter without worrying about sheet existing
            """try:
                with pd.ExcelWriter(full_path, mode="a", engine="openpyxl") as writer:
                    pip_self.eval.test_train_pred.get('train').to_excel(writer, sheet_name=sheet_name, index=False)
            except Exception as e:"""
            print('\nThe excel file could not be exported as .xlsx -> exporting to .paquet:\n',full_path,sheet_name)
            dump(pip_self.eval.test_train_pred.get('train'),full_path.replace('.xlsx','_')+sheet_name+'.paquet')


            print(f"Prediction Values Data exported to {full_path}")
        elif cohort == "test" and "test" in pip_self.eval.test_train_pred and pip_self.eval.test is not None:
            filename = f"Prediction_values_combined.xlsx"
            full_path = os.path.join(combined_output_path, filename)
            if not os.path.exists(full_path):
                wb = Workbook()
                wb.save(full_path)
                wb.close()

            wb = load_workbook(full_path)
            sheet_name = f"{row_subset}_{col_subset}"
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]  # Delete the sheet if it exists
            wb.save(full_path)
            wb.close()

            # Now append using ExcelWriter without worrying about sheet existing
            with pd.ExcelWriter(full_path, mode="a", engine="openpyxl") as writer:
                pip_self.eval.test_train_pred.get('test').reset_index().to_excel(writer, sheet_name=sheet_name, index=False)

            print(f"Prediction Values Data exported to {full_path}")
        else:
            print("Skipping prediction values export (not validation cohort or missing data) - msg1")

    def get_pv_test_train(self, pip_self):
        """Get the predicted values for the training, testing and validation datasets.
        The predicted values are stored in seperated dataframes that are exported and in the eval.val, eval.test and eval.train dictionaries.


        Args:
            pip_self (_type_): _description_

        Returns:
            _type_: _description_
        """

        export_test = pd.DataFrame()
        export_train = pd.DataFrame()
        export_val = None
        if not self.only_val:
            for key, val in pip_self.trained_model.model_with_info.items():
                y_true_train, y_true_test = [], []
                y_pred_train, y_pred_test = [], []
                eid_train, eid_test = [], []
                # iter through all models and get the performance
                y_true_train = y_true_train + list(val["y_train_inner"].status)
                y_pred_train = y_pred_train + list(val["y_pred_train"])
                eid_train = eid_train + list(val["y_train_inner"].index)
                y_true_test = y_true_test + list(val["y_test_inner"].status)
                y_pred_test = y_pred_test + list(val["y_pred_test"])
                eid_test = eid_test + list(val["y_test_inner"].index)

                # export the stats:
                export_train = pd.concat(
                    [export_train, pd.DataFrame({"y_true": y_true_train, "y_pred": y_pred_train, "eid": eid_train})]
                )

                export_test = pd.concat(
                    [export_test, pd.DataFrame({"y_true": y_true_test, "y_pred": y_pred_test, "eid": eid_test})]
                )
            export_train.set_index("eid")
            export_test.set_index("eid")

        if hasattr(pip_self.data, "y_val") and pip_self.data.y_val is not None:
            try:
                export_val = deepcopy(pip_self.data.y_val_orig)
                export_val["y_pred"] = pip_self.master_RFC.predict_proba(
                    pip_self.ohe.transform(pip_self.data.X_val)
                ).tolist()
                self.val.update({"predicted_values": export_val})
            except Exception as e:
                print(f"Could not calculate the prediction values for the validation dataset: {e}")
                export_val = None
        else:
            print("Validation dataset not available.")

        if not self.only_val:
            # merge with the y_orig or y_val_orig to get the additional data from the cancerreg...
            export_train=export_train.merge(pip_self.data.y_orig, left_on='eid', right_index=True, how='left')
            export_test=export_test.merge(pip_self.data.y_orig, left_on='eid', right_index=True, how='left')

            self.test_train_pred.update({"train": export_train, "test": export_test, "val": export_val})
            self.train.update({"predicted_values": export_train})
            self.test.update({"predicted_values": export_test})

        self.val.update({"predicted_values": export_val})
        self.test_train_pred.update({"val": export_val})

        if not hasattr(self, "val") or self.val is None:
            self.val = {}
        self.val["predicted_values"] = export_val

        return export_train, export_test, export_val

    def get_metrics(y_true, y_pred, threshold_steps=0.01):
        """A computational expensive function for the evaluation of a model without taking one specific threshold

        Args:
            y_true (_type_): _description_
            y_pred (_type_): _description_
            threshold_steps (float, optional): _description_. Defaults to 0.01.
        """
